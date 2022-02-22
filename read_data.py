
import openpyxl

class my_dictionary(dict):
  
    def __init__(self):
        self = dict()

    def add(self, key, value):
        self[key] = value

def check_float(potential_float):
    try:
        float(potential_float)
        return True
    except ValueError:
        return False

def get_sheets(file_path):
    try:
        wb_obj = openpyxl.load_workbook(file_path, data_only=True)
        sheets = wb_obj.sheetnames
        # if len(sheets) == 1 and not 0:
        #     sheet_obj = wb_obj.active
        # else:
        #     state, selected_sheet_name = sheet_selection_window(sheets)
        #     sheet_obj = wb_obj[selected_sheet_name]

        return sheets, wb_obj
    except:
        return 'F_Error' , None

def read_sheet_data(sheet_obj):
    # cell_obj = sheet_obj.cell(row = 1, column = 1)
    titles = (sheet_obj.cell(row = 1, column = 1).value, sheet_obj.cell(row = 1, column = 2).value)
    row = sheet_obj.max_row
    # column = sheet_obj.max_column
    data = my_dictionary()
    [data.add(sheet_obj.cell(row = row, column = 1).value, sheet_obj.cell(row = row, column = 2).value) for row in  range (2, row + 1) ]
    
    return(data, titles)

def check_details(values):
    return ((values[0] and values[1] and values[2]) is not None or '') and (values[2].isnumeric())

def check_entered_data(key, value):
    return check_float(value) and (key is not None or '')